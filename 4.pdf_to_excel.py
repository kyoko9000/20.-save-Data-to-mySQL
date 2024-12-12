# import pdfplumber
#
# # Mở file PDF
# with pdfplumber.open('danhsach.pdf') as pdf:
#     for page in pdf.pages:
#         # nếu cần toàn bộ dữ liệu
#         text = page.extract_text()
#         print(text)
#
#         # Nếu cần xử lý các bảng
#         # tables = page.extract_tables()
#         # print(tables)

# -------------------chỉ ghi bảng dữ liệu-------------------
import pdfplumber
import openpyxl

# Mở file PDF
with pdfplumber.open('danhsach.pdf') as pdf:
    # Tạo workbook và worksheet mới
    wb = openpyxl.Workbook()
    ws = wb.active

    for page in pdf.pages:
        # Trích xuất các bảng từ trang
        tables = page.extract_tables()
        # print(tables)
        list = tables[0]

        # Chuyển đổi tất cả các số thành số nguyên
        for row in list[1:]:  # Bỏ qua hàng tiêu đề đầu tiên
            for i in range(len(row)):
                try:
                    row[i] = int(row[i])
                except ValueError:
                    pass  # Nếu không thể chuyển đổi thì bỏ qua

        print(list)

        for row in list:
            ws.append(row)

# Lưu file Excel
wb.save('hoc_sinh.xlsx')

# ------------------ghi luôn dòng tiêu đề và có chỉnh sửa---------------
# import pdfplumber
# import openpyxl
#
# # Mở file PDF
# with pdfplumber.open('danhsach.pdf') as pdf:
#     # Tạo workbook và worksheet mới
#     wb = openpyxl.Workbook()
#     ws = wb.active
#
#     for page in pdf.pages:
#         # Trích xuất các bảng từ trang
#         tables = page.extract_text()
#         # print(tables)
#         # Tách các dòng dữ liệu
#         lines = tables.strip().split("\n")
#         data_list = [line for line in lines]
#         # print(data_list[0])
#
#         tables_1 = page.extract_table()
#         # print(tables_1)
#
#         # Thêm tiêu đề vào đầu danh sách
#         tables_1.insert(0, [data_list[0]])
#
#         # print(tables_1)
#
#         # # Chuyển đổi tất cả các số thành số nguyên
#         for row in tables_1[1:]:  # Bỏ qua hàng tiêu đề đầu tiên
#             for i in range(len(row)):
#                 try:
#                     row[i] = int(row[i])
#                 except ValueError:
#                     pass  # Nếu không thể chuyển đổi thì bỏ qua
#
#         print(tables_1)
#         # lựa chọn 1: Thay thế '\n' bằng khoảng cách
#         for row in tables_1:
#             for i in range(len(row)):
#                 if isinstance(row[i], str):
#                     row[i] = row[i].replace('\n', ' ')
#
#         for row in tables_1:
#             ws.append(row)
#
#         # Nối 5 ô thành 1 ô và chứa text "Bảng Kê Tuyến_Phí Vận Chuyển Theo Loại(Bảng 1)"
#         ws.merge_cells('A1:H1')
#         ws['A1'].value = tables_1[0][0]
#
#         # Định dạng ô hợp nhất
#         ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
#
#         # # lựa chọn 2: Định dạng các ô để ngắt dòng trong ô
#         # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#         #     for cell in row:
#         #         if isinstance(cell.value, str) and '\n' in cell.value:
#         #             cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
#
# # Lưu file Excel
# wb.save('hoc_sinh.xlsx')

