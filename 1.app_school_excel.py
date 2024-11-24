# ************************** man hinh loai 2 *************************
import sys
import openpyxl
# pip install pyqt5
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem
from gui1 import Ui_MainWindow


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.load = None
        self.workbook = None
        self.sheet = None
        self.file_path = None
        self.uic = Ui_MainWindow()
        self.uic.setupUi(self)

        self.uic.pushButton.clicked.connect(self.load_data)
    #     # self.uic.pushButton_3.clicked.connect(self.save_to_excel)
        self.uic.pushButton_4.clicked.connect(self.choose_data)

    def load_data(self):
        # Mở tệp Excel
        self.file_path = 'C:/Users/PC/Desktop/google drive link/1.database_1.xlsx'
        self.workbook = openpyxl.load_workbook(self.file_path)

        # Chọn bảng tính (Sheet)
        self.sheet = self.workbook.active  # Hoặc workbook['Tên_Sheet'] nếu bạn biết tên sheet

        # # Đọc toàn bộ nội dung của tệp Excel
        # for row in sheet.iter_rows(values_only=True):
        #     print(row)

        # Thiết lập kích thước bảng
        self.uic.tableWidget.setRowCount(self.sheet.max_row)
        self.uic.tableWidget.setColumnCount(self.sheet.max_column)

        # Đọc dữ liệu từ Excel và thêm vào QTableWidget
        for i, row in enumerate(self.sheet.iter_rows(values_only=True)):
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.uic.tableWidget.setItem(i, j, item)

        # auto save to excel file after choose data
        self.uic.tableWidget.itemChanged.connect(self.save_to_excel)

    def save_to_excel(self):
        item = self.uic.tableWidget.selectedItems()[0]
        row = item.row()
        column = item.column()
        value = item.text()
        print("row", row + 1, "column", column + 1, "value", value)
        if row == 0:
            print("không được phép thay đổi giá trị hàng này")
        if column == 0 or column == 3:
            # Ghi lại dữ liệu từ QTableWidget vào tệp Excel
            self.sheet.cell(row=row + 1, column=column + 1, value=int(value))
            # Lưu lại tệp Excel
            self.workbook.save(self.file_path)
            print("Dữ liệu dạng số đã được lưu thành công vào tệp Excel.")
        if column == 1 or column == 2 or column == 4:
            # Ghi lại dữ liệu từ QTableWidget vào tệp Excel
            self.sheet.cell(row=row + 1, column=column + 1, value=value)
            # Lưu lại tệp Excel
            self.workbook.save(self.file_path)
            print("Dữ liệu dạng chữ đã được lưu thành công vào tệp Excel.")

    def choose_data(self):
        selected_items = self.uic.tableWidget.selectedItems()
        if selected_items:
            row = selected_items[0].row()
            # data at choose cell
            choose_data = []
            for col in range(self.uic.tableWidget.columnCount()):
                item = self.uic.tableWidget.item(row, col)
                if item:
                    choose_data.append(item.text())
            # print(f"Dữ liệu của hàng {row + 1}: {row_data}")
            # data of header
            header_data = []
            for col in range(self.uic.tableWidget.columnCount()):
                item = self.uic.tableWidget.item(0, col)
                if item:
                    header_data.append(item.text())
            list_data = [header_data, choose_data]
            self.export_to_excel(list_data)
        else:
            print("Không có ô nào được chọn")

    def export_to_excel(self, data):
        # Chuyển đổi giá trị tại vị trí 0 và 3 ra dạng int
        data[1][0] = int(data[1][0])
        data[1][3] = int(data[1][3])
        print("data", data)

        # Lưu dữ liệu vào worksheet
        file_path = 'export_data.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)

        # Lưu workbook vào tệp Excel
        workbook.save(file_path)

        print("Dữ liệu đã được lưu thành công vào tệp Excel.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec())
